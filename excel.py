from openpyxl import Workbook, load_workbook
import os

# Excel setup
file_name = "FitPulse.xlsx"

def saveWorkout(date,count,avgSpeed,duration):

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Date","Reps","Avg Speed (s)","Duration (s)"])

    sheet.append([
        date,
        count,
        avgSpeed,
        duration
    ])

    wb.save(file_name)